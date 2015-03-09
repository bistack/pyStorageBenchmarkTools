#!/usr/bin/python
'''Filename: fioxlsx.py
Author: Sun Zhenyuan <sunzhenyuan@163.com> 2015.03.07. '''

class Lat_percent:
    def __init__(self, msec, percent):
        self.msec = float(msec)
        self.percent = float(percent)

    def __eq__(self, another):
        if self.msec == another.msec:
            return True
        return False

    def __ne__(self, another):
        if self.msec != another.msec:
            return True

        return False

    def __lt__(self, another):
        if self.msec < another.msec:
            return True
        return False

    def __gt__(self, another):
        if self.msec > another.msec:
            return True
        return False

    def add_percent(self, percent):
        self.percent = float(self.percent) + float(percent)

    def get_msec(self):
        return self.msec

    def get_percent(self):
        return self.percent

    def get_msec_percent(self):
        return {str(self.msec):str(self.percent)}

class Fio_result:

    def __init__(self, name):
        self.name = name
        self.latency_percent_list = []         # msec
        self.info_dic = {}
        self.latency_dic = {}     # msec
        self.iops_dic = {}
        self.bandwidth_dic = {}   # MB/s
        self.cpu_usr = 0
        self.cpu_sys = 0

    def __eq__(self, another):
        if self.name == another.name:
            return True
        return False

    def get_name(self):
        return self.name

    def set_info(self, key, info):
        self.info_dic[key] = info

    def get_info(self, key):
        return self.info_dic[key]

    def push_latency_percent(self, msec, percent):
        if float(msec) < float(0.01):
            msec = 0.01
        new = Lat_percent(msec, percent)
        for lat_pct in self.latency_percent_list:
            if lat_pct == new:
                lat_pct.add_percent(percent)
                return

        self.latency_percent_list.append(new)

    def get_latency_percent(self):
        return self.latency_percent_list

def print_err_info(status, info):
    '''print err info'''
    print 'Err: ' + str(status) + ', ' + info

def get_fio_char(fio_result):
    start = fio_result.find('md127')
    end = fio_result.rfind('.txt')

    rw = fio_result.rfind('write')
    if rw < 0:
        rw = fio_result.rfind('read')

    parts = fio_result[start:rw].split('_')
    char = ' '.join(parts[0:2])

    left = fio_result[rw:].find('_')
    parts = fio_result[rw + left + 1:end].split('_')

    task = ' '.join(parts)
    return ' '.join([char, task]).strip()

def test_get_fio_char():
    test = get_fio_char('./md127_7_abc_dev_write_seq_4096.txt')
    if test != 'md127 7 seq 4096':
        print test
        return

    test = get_fio_char('./md127T_6_20150307_123737_read_mds_1.txt')
    if test != 'md127T 6 mds 1':
        print test
        return

test_get_fio_char()

import commands
import sys

def print_fio_result(fio_result_obj):
    rst = fio_result_obj
    print rst.get_name()
    print rst.get_info('Avg Write BW')
    print rst.get_info('Avg Write IOPS')
    print rst.get_info('Avg Write Lat')
    print rst.get_info('Max Write Lat')
    print rst.get_info('Avg Read BW')
    print rst.get_info('Avg Read IOPS')
    print rst.get_info('Avg Read Lat')
    print rst.get_info('Max Read Lat')
    print rst.get_info('Avg CPU Usr')
    print rst.get_info('Avg CPU Sys')
    for lat_pct in rst.get_latency_percent():
        for lat, pct in lat_pct.get_msec_percent().items():
            print '%s %s' %(lat, pct)

def parse_fio_result(fio_result):
    #print fio_result
    (status, fiostr) = commands.getstatusoutput('grep fio-2 ' + fio_result)
    if status or fiostr == None:
        #print_err_info(status, fio_result)
        return

    fio_char = get_fio_char(fio_result)
    rst = Fio_result(fio_char)

    result_fd = file(fio_result, 'r')
    last_type = None

    while True:
        line = result_fd.readline()
        if len(line) == 0:
            break

        iops_start = line.find('iops=')
        if  iops_start >= 0:
            iops_start += len('iops=')
            iops_end = line[iops_start:].find(',')
            iops = line[iops_start:iops_start + iops_end]

            bw_start = line.find('bw=')
            if bw_start >= 0:
                bw_value_start = bw_start + len('bw=')
                bw_end = line[bw_value_start:].find('B')
                bw_value_end = bw_value_start + bw_end - 1
                bw_str = line[bw_value_start: bw_value_end]
                bandwidth = None

                if line[bw_value_end: bw_value_end + 1] == 'K':
                    bandwidth = '%.2f' % (float(bw_str) / 1024)
                elif line[bw_value_end: bw_value_end + 1] == 'M':
                    bandwidth = '%.2f' % (float(bw_str))
                elif line[bw_value_end: bw_value_end + 1] == 'G':
                    bandwidth = '%.2f' % (float(bw_str) * 1024)
                else:
                    bw_str = line[bw_value_start: bw_value_end - 1]
                    bandwidth = '%.2f' % (float(bw_str) / 1024 / 1024)

            if line.find('write') >= 0:
                last_type = 'Write'
                rst.set_info('Avg Write IOPS', iops)
                if bandwidth:
                    rst.set_info('Avg Write BW', bandwidth)
            elif line.find('read') >= 0:
                last_type = 'Read'
                rst.set_info('Avg Read IOPS', iops)
                if bandwidth:
                    rst.set_info('Avg Read BW', bandwidth)

        if line.find('slat') >= 0:
            continue
        elif line.find('clat') >= 0:
            continue

        lat_start = line.find('lat')

        if lat_start >= 0:
            lat_avg_start = line.find('avg=')

            if lat_avg_start >= 0:
                lat_avg_value_start = lat_avg_start + len('avg=')
                lat_avg_end = line[lat_avg_value_start:].find(',')
                lat_avg_value_end = lat_avg_value_start + lat_avg_end
                lat_avg_str = line[lat_avg_value_start: lat_avg_value_end]

                lat_avg = None
                if line.find('usec') >= 0:
                    lat_avg = '%.2f' % (float(lat_avg_str) / 1000)
                elif line.find('msec') >= 0:
                    lat_avg = '%.2f' % (float(lat_avg_str))
                elif line.find('sec') >= 0:
                    lat_avg = '%.2f' % (float(lat_avg_str) * 1000)
                else:
                    continue

                rst.set_info('Avg ' + last_type + ' Lat', lat_avg)

            lat_max_start = line.find('max=')

            if lat_max_start >= 0:
                lat_max_value_start = lat_max_start + len('max=')
                lat_max_end = 0
                lat_max_end_K = line[lat_max_value_start:].find('K')
                lat_max_end_M = line[lat_max_value_start:].find('M')
                if lat_max_end_K >= 0:
                    lat_max_end = lat_max_end_K
                elif lat_max_end_M >= 0:
                    lat_max_end = lat_max_end_M
                else:
                    lat_max_end = line[lat_max_value_start:].find(',')

                lat_max_value_end = lat_max_value_start + lat_max_end
                lat_max_str = line[lat_max_value_start: lat_max_value_end]

                lat_max = None

                if lat_max_end_K >= 0:
                    lat_max = '%.2f' % (float(lat_max_str) * 1000)
                elif lat_max_end_M >= 0:
                    lat_max = '%.2f' % (float(lat_max_str) * 1000 * 1000)
                else:
                    lat_max = '%.2f' % (float(lat_max_str))

                if line.find('usec') >= 0:
                    lat_max = '%.2f' % (float(lat_max) / 1000)
                elif line.find('msec') >= 0:
                    lat_max = '%.2f' % (float(lat_max))
                elif line.find('sec') >= 0:
                    lat_max = '%.2f' % (float(lat_max) * 1000)
                else:
                    continue

                rst.set_info('Max ' + last_type + ' Lat', lat_max)

            if line.find('%') >= 0:
                base = 1
                if line.find('usec') >= 0:
                    base = 1
                elif line.find('msec') >= 0:
                    base = 1000
                elif line.find('sec') >= 0:
                    base = 1000 * 1000
                else:
                    continue

                v_start = line.find(':')
                lat_pcts = line[v_start + 1:].replace(',', '').split()
                for part in lat_pcts:
                    lat_end = part.rfind('=')
                    lat_start = 0
                    pre = part.find('=')

                    if lat_end != pre:
                        lat_start = pre + 1

                    pct_end = part.find('%')
                    lat_str = part[lat_start:lat_end]
                    lat_value = '%.2f' % (float(lat_str) * base / 1000)
                    pct_str = part[lat_end+1:pct_end]
                    pct_value = '%.2f' % (float(pct_str))
                    rst.push_latency_percent(lat_value, pct_value)

        if line.find('cpu') >= 0:
            v_start = line.find(':')
            lat_pcts = line[v_start + 1:].replace(',', '').split()

            for part in lat_pcts:
                cpu_start = 0
                cpu_end = part.rfind('=')
                pct_end = part.find('%')
                if pct_end < 0:
                    break

                cpu_str = part[cpu_start:cpu_end]
                pct_str = part[cpu_end+1:pct_end]
                pct_value = '%.2f' % (float(pct_str))

                if cpu_str == 'usr':
                    rst.set_info('Avg CPU Usr', pct_value)
                elif cpu_str == 'sys':
                    rst.set_info('Avg CPU Sys', pct_value)

    result_fd.close()
    return rst


from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl import utils
import os

info_names = ['Avg Write BW', 'Avg Write IOPS', 'Avg Write Lat', 'Max Write Lat',
              'Avg Read BW', 'Avg Read IOPS', 'Avg Read Lat', 'Max Read Lat',
              'Avg CPU Usr',  'Avg CPU Sys']

def get_compare_space():
    return 10 - len(info_names) % 10

def get_excel_workbook(excel_name):
    wb = None
    try:
        wb = load_workbook(excel_name + '.xlsx')
    except:
        if not wb:
            print 'Excel file %s.xlsx not found, create one' % (excel_name)
            wb = Workbook()

    return wb

def excel_init_worksheet(ws, test_type, offset):
    rows_names = ['Test ' + test_type] + info_names

    row_id = offset
    for row_name in rows_names:
        cell = 'A' + str(row_id)
        ws[cell] = row_name
        row_id += 1

def get_fio_dev_conf(fio_result_name):
    parts = fio_result_name.split()
    if parts[2] == 'seq':
        return ' '.join(parts[0:3])
    else:
        return ' '.join([parts[0]] + parts[2:])

def get_excel_worksheet(fio_result_name, wb):
    ws_name = get_fio_dev_conf(fio_result_name).replace('T', '')
    ws = None

    try:
        if wb:
            ws = wb.get_sheet_by_name(ws_name)
    except:
        if not ws:
            ws = wb.create_sheet()
            ws.title = ws_name

    return ws

def get_column_name(fio_result_name):
    parts = fio_result_name.split()
    if parts[2] == 'seq':
        return int(parts[3])
    else:
        return int(parts[1])

def search_cell_in_row(ws, row, col_label, l_col):
    last_col_letter = utils.get_column_letter(l_col)

    cell_found = None
    for arow in ws.iter_rows('A' + str(row) + ':' + last_col_letter + str(row)):
        for cell in arow:
            if cell.value == col_label:
                cell_found = cell
                break
    return cell_found

def move_back_one_column(ws, col):
    l_row = len(info_names) * 2  + 2 + get_compare_space()

    if (l_row < 1):
        print 'Err: highest row %d' % (str(l_row))
        return

    for row in range(1, l_row):
        i_cell = ws.cell(row=row, column=col)
        if not i_cell.value:
            continue

        next_cell = ws.cell(row=row, column=col+1)
        next_cell.value = i_cell.value
        i_cell.value = None

def move_back_columns(ws, begin_col, end_col):
    for col_i in range(end_col, begin_col, -1):
        move_back_one_column(ws, col_i)

def get_pos_insert_sort_one_column(ws, col_label, l_col):
    l_col_letter = utils.get_column_letter(l_col)

    found_cell = None

    # alwasy see the first row for column's label, column A is label
    for arow in ws.iter_rows('B1:' + l_col_letter + '1'):
        for cell in arow:
            if not cell.value or int(cell.value) > int(col_label):
                found_cell = cell
                break

    if not found_cell:
        col = l_col + 1
        col_letter = utils.get_column_letter(col)
    else:
        col_letter = found_cell.column
        if found_cell.value:
            col_idx = utils.column_index_from_string(found_cell.column)
            move_back_columns(ws, col_idx, l_col)

    return col_letter

def excel_add_col_label_in_row(ws, row, col_label, l_col):

    # always search the first row
    cell_found = search_cell_in_row(ws, 1, col_label, l_col)

    if cell_found:
        #print 'found %s, column: %s' % (cell_found.value, cell_found.column)
        col_letter = cell_found.column
    else:
        col_letter = get_pos_insert_sort_one_column(ws, col_label, l_col)

    cell = ws.cell(col_letter + str(row))
    cell.value = col_label

    return col_letter

def excel_add_col_label_in_compare_rows(ws, row, col_label, l_col):
    col_letter_a = excel_add_col_label_in_row(ws, 1, col_label, l_col)

    if row > 1:
        col_letter_b = excel_add_col_label_in_row(ws, row, col_label, l_col)

        if col_letter_a != col_letter_b:
            print ('Err: column letter not match, %s %s'
                   % (col_letter_a, col_letter_b))
            return None

    return col_letter_a

def get_start_row(fio_result_name):
    conf = get_fio_dev_conf(fio_result_name)
    parts = conf.split()
    start_row = 1
    
    dev_type = 'T'
    if parts[0].find('T') < 0:
        space = get_compare_space()
        start_row += len(info_names) + space
        dev_type = 'MD'

    return (dev_type, start_row)

def check_start_row_label(ws, row):
    row_label = ws.cell('A' + str(row)).value

    if  row_label.find('Test') < 0:
        print 'Err: compare row %s' % (row_label)
        return 0
    
    return 1

def excel_add_fio_result(fio_result_obj, wb):
    ws = get_excel_worksheet(fio_result_obj.name, wb)
    if not wb or not ws:
        return

    (dev_type, start_row) = get_start_row(fio_result_obj.name)
    excel_init_worksheet(ws, dev_type, start_row)
    
    if not check_start_row_label(ws, start_row):
        return

    l_col = ws.get_highest_column()
    col_label = get_column_name(fio_result_obj.name)
    col_letter = excel_add_col_label_in_compare_rows(ws, start_row, col_label,
                                                     l_col)

    if not col_letter:
        return

    i_row = start_row + 1
    for name in info_names:
        if fio_result_obj.info_dic.has_key(name):
            value = fio_result_obj.get_info(name)
            cell = ws.cell(col_letter + str(i_row))
            if cell.value:
                label_cell = ws.cell('A' + str(i_row))
                if (label_cell.value == 'Avg CPU Usr' or
                    label_cell.value == 'Avg CPU Sys'):
                    if float(value) > float(cell.value):
                        cell.value = float(value)

                elif (float(cell.value) > float(value) * 2 or
                    float(value) > float(cell.value) * 2):
                    row_label = ws.cell('A' + str(cell.row)).value
                    print (('sheet: %s, col_label: %s, dev type: %s, ' +
                           'label: %s, old: %s, new: %s')
                           % (ws.title, col_label, dev_type, str(row_label), 
                              cell.value, value))
                else:
                    avg = '%.2f' % (float(cell.value) + float(value) / 2)
                    cell.value = float(avg)
            else:
                cell.value =  float(value)

        i_row +=1

def get_file_list(idir):
    '''files in a dir'''
    (status, files) = commands.getstatusoutput('ls ' + idir)
    if (status):
        print_err_info(status, files)
        return

    return files.split()

def get_comparable_row_idx_by_label(ws, row_label):
    # we skip the first label with diff dev type 
    idx1 = 2
    for row_idx in range(2, ws.get_highest_row() / 2):
        cell = ws.cell('A' + str(row_idx))
        if cell and cell.value == row_label:
            idx1 = cell.row
            break

    space = get_compare_space()
    return (idx1, idx1 + space + len(info_names))

def check_comparalbe_row_idxes(ws, row1, row2):
    cell1 = ws.cell('A' + str(row1))
    cell2 = ws.cell('A' + str(row2))

    if cell1 and cell2 and (cell1.value == cell2.value):
        return 1

    print ('Warn: row %d: %s  %d: %s labels not match'
           % (row1, cell1.value, row2, cell2.value))
    return 0

from openpyxl.charts import ScatterChart, Reference, Series

def excel_draw_compare_chart(ws, row_label):
    '''Warning: Openpyxl currently supports chart creation within a worksheet only.
    Charts in existing workbooks will be lost'''

    (row1, row2) = get_comparable_row_idx_by_label(ws, row_label)

    if not check_comparalbe_row_idxes(ws, row1, row2):
        return

    start_col = 2
    end_col = ws.get_highest_column()

    xvalues = Reference(ws, (1, start_col), (1, end_col))

    values1 = Reference(ws, (row1, start_col), (row1, end_col))
    series1 = Series(values1, title=row_label + ' T', xvalues=xvalues)

    values2 = Reference(ws, (row2, start_col), (row2, end_col))
    series2 = Series(values2, title=row_label + ' MD', xvalues=xvalues)

    lines = ScatterChart()
    lines.append(series1)
    lines.append(series2)
    ws.add_chart(lines)

def excel_draw_all_char(wb):
    for ws_name in wb.sheetnames:
        ws = wb.get_sheet_by_name(ws_name)
        if wb.get_index(ws) == 0:
            continue

        for label in info_names:
            excel_draw_compare_chart(ws, label)

import time
def parse_all_test_file():
    excel_file_name = 'test_result_' + time.strftime('%Y%m%d_%H%M%S')

    if len(sys.argv) > 0:
        rslt_dir = sys.argv[1]
        wb = get_excel_workbook(excel_file_name)
        files = get_file_list(rslt_dir)

        if files and len(files) > 0:
            for afile in files:
                file_p = rslt_dir + os.sep + afile

                matched = commands.getoutput('echo ' + afile +
                                             '| grep -e write[0-9]')
                if matched:
                    fparts = afile.split('write')
                    newfile = 'write_seq_'.join(fparts)
                    nfile_p = rslt_dir + os.sep + newfile
                    print nfile_p
                    commands.getoutput('mv ' + file_p + ' ' + nfile_p)

                matched = commands.getoutput('echo ' + afile +
                                             '| grep -e read[0-9]')

                if matched:
                    fparts = afile.split('read')
                    newfile = 'read_seq_'.join(fparts)
                    nfile_p = rslt_dir + os.sep + newfile
                    print nfile_p
                    commands.getoutput('mv ' + file_p + ' ' + nfile_p)

                rst = parse_fio_result(file_p)
                if rst:
                    excel_add_fio_result(rst, wb)

        excel_draw_all_char(wb)
        wb.save(excel_file_name + '.xlsx')

parse_all_test_file()
